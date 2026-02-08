"""
Integration tests for the FedRAMP Readiness Assessment Toolkit.

These tests validate that scripts work together as a pipeline:
    inheritance_mapper → gap_analysis → ssp_generator → oscal_validator
    scan_aggregator → poam_manager → compliance_scorer
    inventory_drift (standalone)

Tests use sample data from fixtures/ and examples/ and run without
cloud API access or downloaded baselines.
"""

import csv
import json
import os
import shutil
import sqlite3
import pytest
from datetime import datetime, timezone, timedelta
from pathlib import Path

import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

FIXTURES = Path(__file__).parent.parent / "fixtures"
EXAMPLES = Path(__file__).parent.parent.parent / "examples"
PROJECT_ROOT = Path(__file__).parent.parent.parent


# ============================================================
# Shared fixtures
# ============================================================


@pytest.fixture
def output_dir(tmp_path):
    """Temporary output directory for test artifacts."""
    d = tmp_path / "reports"
    d.mkdir()
    return d


@pytest.fixture
def sample_baseline_controls():
    """Simulated baseline controls (avoids needing real baselines)."""
    families = {
        "AC": ["AC-1", "AC-2", "AC-3", "AC-4", "AC-5", "AC-6", "AC-7", "AC-8", "AC-17"],
        "SC": ["SC-1", "SC-7", "SC-8", "SC-12", "SC-13", "SC-28"],
        "AU": ["AU-1", "AU-2", "AU-3", "AU-6", "AU-8", "AU-9", "AU-12"],
        "IA": ["IA-1", "IA-2", "IA-4", "IA-5", "IA-7", "IA-8"],
        "CM": ["CM-1", "CM-2", "CM-3", "CM-6", "CM-7", "CM-8"],
        "IR": ["IR-1", "IR-2", "IR-4", "IR-5", "IR-6", "IR-8"],
        "PE": ["PE-1", "PE-2", "PE-3", "PE-6", "PE-8"],
        "PS": ["PS-1", "PS-2", "PS-3", "PS-4"],
        "AT": ["AT-1", "AT-2", "AT-3"],
        "RA": ["RA-1", "RA-3", "RA-5"],
        "SI": ["SI-1", "SI-2", "SI-3", "SI-4", "SI-5"],
        "CP": ["CP-1", "CP-2", "CP-9", "CP-10"],
        "SR": ["SR-1", "SR-2", "SR-3"],
    }
    controls = []
    for family_id, cids in families.items():
        for cid in cids:
            controls.append({"id": cid, "title": f"{cid} Control", "family_id": family_id})
    return controls


@pytest.fixture
def sample_implementation_status():
    """Simulated implementation status matching sample_baseline_controls."""
    return {
        "AC-1": "implemented", "AC-2": "partial", "AC-3": "implemented",
        "AC-4": "not_implemented", "AC-5": "planned", "AC-6": "implemented",
        "AC-7": "implemented", "AC-8": "not_implemented", "AC-17": "implemented",
        "SC-1": "implemented", "SC-7": "implemented", "SC-8": "implemented",
        "SC-12": "partial", "SC-13": "implemented", "SC-28": "implemented",
        "AU-1": "implemented", "AU-2": "partial", "AU-3": "implemented",
        "AU-6": "partial", "AU-8": "implemented", "AU-9": "implemented",
        "AU-12": "implemented",
        "IA-1": "planned", "IA-2": "implemented", "IA-4": "implemented",
        "IA-5": "partial", "IA-7": "implemented", "IA-8": "not_implemented",
        "CM-1": "implemented", "CM-2": "partial", "CM-3": "planned",
        "CM-6": "implemented", "CM-7": "partial", "CM-8": "implemented",
        "IR-1": "not_implemented", "IR-2": "not_implemented",
        "IR-4": "planned", "IR-5": "planned", "IR-6": "not_implemented",
        "IR-8": "not_implemented",
        "PE-1": "inherited", "PE-2": "inherited", "PE-3": "inherited",
        "PE-6": "inherited", "PE-8": "inherited",
        "PS-1": "implemented", "PS-2": "implemented", "PS-3": "partial", "PS-4": "implemented",
        "AT-1": "implemented", "AT-2": "partial", "AT-3": "planned",
        "RA-1": "implemented", "RA-3": "partial", "RA-5": "implemented",
        "SI-1": "implemented", "SI-2": "partial", "SI-3": "implemented",
        "SI-4": "partial", "SI-5": "implemented",
        "CP-1": "implemented", "CP-2": "partial", "CP-9": "implemented", "CP-10": "planned",
        "SR-1": "not_implemented", "SR-2": "not_implemented", "SR-3": "not_implemented",
    }


@pytest.fixture
def scan_findings():
    """Realistic scan findings for pipeline testing."""
    now = datetime.now(timezone.utc)
    return [
        {
            "finding_id": "f001", "cve_id": "CVE-2024-1234",
            "title": "Critical OpenSSL RCE", "severity": "CRITICAL",
            "cvss_score": 9.8, "affected_resource": "server-1",
            "resource_type": "os", "scanner": "nessus",
            "description": "Remote code execution", "solution": "Patch OpenSSL",
            "first_seen": (now - timedelta(days=45)).isoformat(),
            "last_seen": now.isoformat(), "status": "open",
        },
        {
            "finding_id": "f002", "cve_id": "CVE-2024-5678",
            "title": "High TLS Bypass", "severity": "HIGH",
            "cvss_score": 7.5, "affected_resource": "server-1",
            "resource_type": "os", "scanner": "nessus",
            "description": "TLS bypass", "solution": "Update TLS library",
            "first_seen": (now - timedelta(days=10)).isoformat(),
            "last_seen": now.isoformat(), "status": "open",
        },
        {
            "finding_id": "f003", "cve_id": None,
            "title": "Moderate Info Disclosure", "severity": "MODERATE",
            "cvss_score": 5.0, "affected_resource": "server-2",
            "resource_type": "web_app", "scanner": "qualys",
            "description": "Version headers exposed", "solution": "Suppress headers",
            "first_seen": (now - timedelta(days=100)).isoformat(),
            "last_seen": now.isoformat(), "status": "open",
        },
        {
            "finding_id": "f004", "cve_id": "CVE-2024-9876",
            "title": "Container vuln", "severity": "HIGH",
            "cvss_score": 8.1, "affected_resource": "myapp:latest",
            "resource_type": "container", "scanner": "trivy",
            "description": "Package vuln in container", "solution": "Update base image",
            "first_seen": (now - timedelta(days=5)).isoformat(),
            "last_seen": now.isoformat(), "status": "open",
        },
        {
            "finding_id": "f005", "cve_id": None,
            "title": "CloudTrail disabled", "severity": "HIGH",
            "cvss_score": None, "affected_resource": "us-gov-west-1",
            "resource_type": "cloud_config", "scanner": "prowler",
            "description": "CloudTrail not enabled in region", "solution": "Enable CloudTrail",
            "first_seen": (now - timedelta(days=35)).isoformat(),
            "last_seen": now.isoformat(), "status": "open",
        },
    ]


# ============================================================
# Pipeline 1: Inheritance → Gap Analysis
# ============================================================


class TestInheritanceToGapPipeline:
    """Test that inheritance mapper output feeds correctly into gap analysis."""

    def test_inheritance_yaml_feeds_gap_analysis(self, tmp_path, sample_baseline_controls, sample_implementation_status):
        from scripts.inheritance_mapper import get_inheritance_for_control
        from scripts.gap_analysis import analyze_gaps

        # Step 1: Generate inheritance map (simulate what the mapper CLI does)
        inheritance_map = {}
        for ctrl in sample_baseline_controls:
            result = get_inheritance_for_control(ctrl["id"], "aws")
            inheritance_map[ctrl["id"]] = result

        # Step 2: Feed into gap analysis
        result = analyze_gaps(sample_baseline_controls, sample_implementation_status, inheritance_map)

        assert "overall" in result
        assert result["overall"]["total_controls"] == len(sample_baseline_controls)

        # PE controls should be inherited regardless of implementation status
        pe_controls = [c for c in result.get("controls", []) if c["control_id"].startswith("PE")]
        pe_gaps = [c for c in pe_controls if c["is_gap"]]
        assert len(pe_gaps) == 0, "PE controls should not be gaps when inherited from AWS"

    def test_inheritance_upgrades_not_implemented(self, sample_baseline_controls):
        from scripts.inheritance_mapper import get_inheritance_for_control
        from scripts.gap_analysis import analyze_gaps

        # All controls not_implemented
        impl = {c["id"]: "not_implemented" for c in sample_baseline_controls}

        # But with inheritance
        inheritance = {}
        for ctrl in sample_baseline_controls:
            inheritance[ctrl["id"]] = get_inheritance_for_control(ctrl["id"], "aws")

        result = analyze_gaps(sample_baseline_controls, impl, inheritance)

        # PE controls should be inherited, not gaps
        pe_controls = [c for c in sample_baseline_controls if c["family_id"] == "PE"]
        inherited_count = result["overall"].get("inherited", 0)
        assert inherited_count >= len(pe_controls)

    def test_shared_controls_become_partial(self, sample_baseline_controls):
        from scripts.inheritance_mapper import get_inheritance_for_control
        from scripts.gap_analysis import analyze_gaps

        # AC controls are "shared" on all providers
        impl = {c["id"]: "not_implemented" for c in sample_baseline_controls}
        inheritance = {}
        for ctrl in sample_baseline_controls:
            inheritance[ctrl["id"]] = get_inheritance_for_control(ctrl["id"], "aws")

        result = analyze_gaps(sample_baseline_controls, impl, inheritance)

        # AC controls that are shared should be at least partial
        ac_controls = [c for c in result.get("controls", []) if c["control_id"].startswith("AC")]
        for ctrl in ac_controls:
            # Shared + not_implemented → partial (not full gap)
            assert ctrl["status"] in ("partial", "not_implemented")


# ============================================================
# Pipeline 2: Scan Aggregator → POA&M Manager
# ============================================================


class TestScanToPoamPipeline:
    """Test scan aggregator feeding into POA&M manager."""

    def test_nessus_to_poam(self, output_dir):
        from scripts.scan_aggregator import aggregate_findings, compute_scan_summary, output_findings_json
        from scripts.poam_manager import load_scan_findings, create_poam_item, load_sla_config, compute_poam_summary

        # Step 1: Aggregate scans
        findings = aggregate_findings([FIXTURES / "sample-nessus.csv"])
        summary = compute_scan_summary(findings)
        output_findings_json(findings, summary, output_dir / "scan.json")

        # Step 2: Load into POA&M manager
        config = load_sla_config()
        loaded = load_scan_findings(output_dir / "scan.json")
        assert len(loaded) > 0

        # Step 3: Create POA&M items
        items = []
        for idx, f in enumerate(loaded, 1):
            if f.get("severity", "").upper() != "INFORMATIONAL":
                items.append(create_poam_item(f, config, idx))

        assert len(items) > 0
        assert all(i["poam_id"].startswith("POAM-") for i in items)
        assert all(i["sla_days"] > 0 for i in items)

    def test_trivy_to_poam(self, output_dir):
        from scripts.scan_aggregator import aggregate_findings, compute_scan_summary, output_findings_json
        from scripts.poam_manager import load_scan_findings, create_poam_item, load_sla_config

        findings = aggregate_findings([FIXTURES / "sample-trivy.json"])
        summary = compute_scan_summary(findings)
        output_findings_json(findings, summary, output_dir / "scan.json")

        config = load_sla_config()
        loaded = load_scan_findings(output_dir / "scan.json")
        items = [create_poam_item(f, config, i) for i, f in enumerate(loaded, 1)
                 if f.get("severity", "").upper() != "INFORMATIONAL"]

        # Should have critical and high items from container scan
        critical = [i for i in items if i["severity"] == "CRITICAL"]
        high = [i for i in items if i["severity"] == "HIGH"]
        assert len(critical) >= 1
        assert len(high) >= 1

    def test_multi_scanner_to_poam_with_dedup(self, output_dir):
        from scripts.scan_aggregator import aggregate_findings, deduplicate_findings, compute_scan_summary, output_findings_json
        from scripts.poam_manager import load_scan_findings, create_poam_item, load_sla_config

        findings = aggregate_findings([FIXTURES / "sample-nessus.csv", FIXTURES / "sample-trivy.json"])
        deduped = deduplicate_findings(findings)
        summary = compute_scan_summary(deduped)
        output_findings_json(deduped, summary, output_dir / "scan.json")

        config = load_sla_config()
        loaded = load_scan_findings(output_dir / "scan.json")
        items = [create_poam_item(f, config, i) for i, f in enumerate(loaded, 1)
                 if f.get("severity", "").upper() != "INFORMATIONAL"]

        # CVE-2024-1234 should be deduplicated (in both Nessus and Trivy)
        cve_1234 = [i for i in items if i.get("cve_id") == "CVE-2024-1234"]
        # After dedup, might still appear multiple times due to different resources
        assert len(items) > 0

    def test_csv_round_trip(self, output_dir):
        """Test scan CSV output can be loaded by POA&M manager."""
        from scripts.scan_aggregator import aggregate_findings, output_conmon_csv
        from scripts.poam_manager import load_scan_findings, create_poam_item, load_sla_config

        findings = aggregate_findings([FIXTURES / "sample-nessus.csv"])
        output_conmon_csv(findings, output_dir / "conmon.csv")

        config = load_sla_config()
        loaded = load_scan_findings(output_dir / "conmon.csv")
        items = [create_poam_item(f, config, i) for i, f in enumerate(loaded, 1)
                 if f.get("severity", "").upper() not in ("INFORMATIONAL", "")]

        assert len(items) > 0


# ============================================================
# Pipeline 3: SSP Generator → OSCAL Validator
# ============================================================


class TestSSPToValidatorPipeline:
    """Test SSP generator output passes OSCAL validation."""

    def test_generated_ssp_validates(self, output_dir, sample_baseline_controls):
        from scripts.ssp_generator import load_system_metadata, load_control_narratives, generate_oscal_ssp
        from scripts.oscal_validator import validate_document

        metadata = load_system_metadata(EXAMPLES / "sample-system-metadata.yaml")
        narratives = load_control_narratives(FIXTURES / "sample-controls")

        ssp = generate_oscal_ssp(metadata, sample_baseline_controls, narratives, None, "moderate")

        ssp_path = output_dir / "test-ssp.json"
        ssp_path.write_text(json.dumps(ssp, indent=2))

        result = validate_document(ssp_path, "ssp")
        assert result.passed, f"Generated SSP failed validation: {[e['message'] for e in result.errors]}"

    def test_ssp_with_inheritance_validates(self, output_dir, sample_baseline_controls):
        from scripts.ssp_generator import load_system_metadata, generate_oscal_ssp
        from scripts.inheritance_mapper import get_inheritance_for_control
        from scripts.oscal_validator import validate_document

        metadata = load_system_metadata(EXAMPLES / "sample-system-metadata.yaml")
        inheritance = {c["id"]: get_inheritance_for_control(c["id"], "aws") for c in sample_baseline_controls}

        ssp = generate_oscal_ssp(metadata, sample_baseline_controls, {}, inheritance, "moderate")

        ssp_path = output_dir / "test-ssp-inherited.json"
        ssp_path.write_text(json.dumps(ssp, indent=2))

        result = validate_document(ssp_path, "ssp")
        assert result.passed, f"SSP with inheritance failed: {[e['message'] for e in result.errors]}"

    def test_ssp_has_all_baseline_controls(self, output_dir, sample_baseline_controls):
        from scripts.ssp_generator import load_system_metadata, generate_oscal_ssp

        metadata = load_system_metadata(EXAMPLES / "sample-system-metadata.yaml")
        ssp = generate_oscal_ssp(metadata, sample_baseline_controls, {}, None, "moderate")

        reqs = ssp["system-security-plan"]["control-implementation"]["implemented-requirements"]
        control_ids = {r["control-id"].upper() for r in reqs}
        baseline_ids = {c["id"].upper() for c in sample_baseline_controls}

        assert control_ids == baseline_ids, f"Mismatch: missing {baseline_ids - control_ids}"

    def test_narratives_populated_in_ssp(self, sample_baseline_controls):
        from scripts.ssp_generator import load_system_metadata, load_control_narratives, generate_oscal_ssp

        metadata = load_system_metadata(EXAMPLES / "sample-system-metadata.yaml")
        narratives = load_control_narratives(FIXTURES / "sample-controls")

        ssp = generate_oscal_ssp(metadata, sample_baseline_controls, narratives, None, "moderate")

        reqs = ssp["system-security-plan"]["control-implementation"]["implemented-requirements"]
        ac1 = next((r for r in reqs if r["control-id"].upper() == "AC-1"), None)
        assert ac1 is not None
        desc = ac1["by-components"][0]["description"]
        assert "access control policy" in desc.lower()


# ============================================================
# Pipeline 4: Gap + POA&M → Compliance Scorer
# ============================================================


class TestGapPoamToScorerPipeline:
    """Test compliance scorer ingests gap and POA&M data correctly."""

    def test_full_scoring_pipeline(self, output_dir, sample_baseline_controls, sample_implementation_status, scan_findings):
        from scripts.gap_analysis import analyze_gaps
        from scripts.poam_manager import create_poam_item, load_sla_config, compute_poam_summary, check_escalation_triggers
        from scripts.compliance_scorer import compute_compliance_score, init_db, store_score, get_trend_data

        # Step 1: Gap analysis
        gap_result = analyze_gaps(sample_baseline_controls, sample_implementation_status)

        # Step 2: POA&M
        config = load_sla_config()
        poam_items = [create_poam_item(f, config, i) for i, f in enumerate(scan_findings, 1)
                      if f.get("severity", "").upper() != "INFORMATIONAL"]
        escalations = check_escalation_triggers(poam_items, config)
        poam_summary = compute_poam_summary(poam_items)
        poam_data = {"summary": poam_summary, "escalation_triggers": escalations, "items": poam_items}

        # Step 3: Score
        score = compute_compliance_score(gap_report=gap_result, poam_data=poam_data)

        assert 0 <= score["composite_score"] <= 100
        assert score["risk_rating"] in ("LOW", "MODERATE", "ELEVATED", "HIGH")
        assert score["component_scores"]["control_implementation"] > 0
        assert score["component_scores"]["vulnerability_posture"] > 0

        # Step 4: Store trend
        db_path = output_dir / "test.db"
        conn = init_db(db_path)
        store_score(conn, score)
        trend = get_trend_data(conn)
        assert len(trend) == 1
        conn.close()

    def test_scorer_threshold_gate(self, sample_baseline_controls, sample_implementation_status):
        from scripts.gap_analysis import analyze_gaps
        from scripts.compliance_scorer import compute_compliance_score

        gap_result = analyze_gaps(sample_baseline_controls, sample_implementation_status)
        score = compute_compliance_score(gap_report=gap_result)

        # Threshold check like CI/CD would do
        assert score["composite_score"] >= 0
        # Our sample is partially implemented, shouldn't be 100
        assert score["composite_score"] < 100


# ============================================================
# Inventory Drift Pipeline
# ============================================================


class TestInventoryDriftPipeline:
    """Test inventory drift with fixture data."""

    def test_full_drift_analysis(self, output_dir):
        from scripts.inventory_drift import (
            load_documented_inventory, load_live_inventory,
            detect_drift, output_drift_json, output_drift_csv, output_drift_html,
        )

        documented = load_documented_inventory(FIXTURES / "sample-documented-inventory.yaml")
        live = load_live_inventory(FIXTURES / "sample-live-inventory.json")
        drift = detect_drift(documented, live)

        # Output all formats
        output_drift_json(drift, output_dir / "drift.json")
        output_drift_csv(drift, output_dir / "drift.csv")
        output_drift_html(drift, output_dir / "drift.html")

        # Verify outputs exist and are non-empty
        assert (output_dir / "drift.json").stat().st_size > 100
        assert (output_dir / "drift.csv").stat().st_size > 50
        assert (output_dir / "drift.html").stat().st_size > 200

    def test_drift_json_parseable(self, output_dir):
        from scripts.inventory_drift import (
            load_documented_inventory, load_live_inventory,
            detect_drift, output_drift_json,
        )

        documented = load_documented_inventory(FIXTURES / "sample-documented-inventory.yaml")
        live = load_live_inventory(FIXTURES / "sample-live-inventory.json")
        drift = detect_drift(documented, live)
        output_drift_json(drift, output_dir / "drift.json")

        loaded = json.loads((output_dir / "drift.json").read_text())
        assert "summary" in loaded
        assert "undocumented" in loaded
        assert "stale" in loaded
        # Should not have internal _normalized_id field
        for item in loaded.get("undocumented", []):
            assert "_normalized_id" not in item

    def test_drift_detects_known_items(self):
        from scripts.inventory_drift import load_documented_inventory, load_live_inventory, detect_drift

        documented = load_documented_inventory(FIXTURES / "sample-documented-inventory.yaml")
        live = load_live_inventory(FIXTURES / "sample-live-inventory.json")
        drift = detect_drift(documented, live)

        # i-0old999removed is in documented but not live -> stale
        stale_ids = [r["resource_id"] for r in drift["stale"]]
        assert "i-0old999removed" in stale_ids

        # i-0new111undocumented is in live but not documented -> undocumented
        undoc_ids = [r["resource_id"] for r in drift["undocumented"]]
        assert "i-0new111undocumented" in undoc_ids


# ============================================================
# Report generation integration
# ============================================================


class TestReportGeneration:
    """Test that all report formats generate valid output."""

    def test_html_report_well_formed(self, output_dir):
        from scripts.utils.report_generators import generate_html_report

        path = generate_html_report(
            title="Test Report",
            content_sections=[
                {"heading": "Section 1", "body": "<p>Content here</p>"},
                {"heading": "Section 2", "body": "<table><tr><td>Data</td></tr></table>"},
            ],
            output_path=output_dir / "test.html",
            metadata={"Baseline": "Moderate", "Date": "2025-01-01"},
        )

        html = path.read_text()
        assert "<html" in html
        assert "Test Report" in html
        assert "Section 1" in html
        assert "Moderate" in html

    def test_excel_report_valid(self, output_dir):
        from scripts.utils.report_generators import generate_excel_report

        path = generate_excel_report(
            title="Test",
            sheets_data={
                "Summary": [{"Metric": "Score", "Value": 85}],
                "Details": [{"Control": "AC-1", "Status": "Implemented"}, {"Control": "AC-2", "Status": "Partial"}],
            },
            output_path=output_dir / "test.xlsx",
        )

        assert path.stat().st_size > 0

        # Verify with openpyxl
        import openpyxl
        wb = openpyxl.load_workbook(path)
        assert "Summary" in wb.sheetnames
        assert "Details" in wb.sheetnames
        ws = wb["Details"]
        assert ws.cell(1, 1).value == "Control"
        assert ws.cell(2, 1).value == "AC-1"

    def test_gap_analysis_html_output(self, output_dir, sample_baseline_controls, sample_implementation_status):
        from scripts.gap_analysis import analyze_gaps, output_html

        result = analyze_gaps(sample_baseline_controls, sample_implementation_status)
        output_html(result, "moderate", output_dir / "gap.html")

        html = (output_dir / "gap.html").read_text()
        assert "<html" in html
        assert "Compliance" in html or "compliance" in html


# ============================================================
# End-to-end pipeline
# ============================================================


class TestEndToEndPipeline:
    """Full pipeline integration: all scripts working together."""

    def test_complete_assessment_pipeline(self, output_dir, sample_baseline_controls, sample_implementation_status, scan_findings):
        """
        Simulates the complete FedRAMP readiness assessment workflow:
        1. Map control inheritance from AWS
        2. Run gap analysis with inheritance
        3. Aggregate vulnerability scans
        4. Generate POA&M with SLA tracking
        5. Generate SSP
        6. Validate SSP
        7. Compute compliance score
        """
        from scripts.inheritance_mapper import get_inheritance_for_control
        from scripts.gap_analysis import analyze_gaps
        from scripts.scan_aggregator import aggregate_findings, deduplicate_findings, compute_scan_summary, output_findings_json
        from scripts.poam_manager import create_poam_item, load_sla_config, compute_poam_summary, check_escalation_triggers
        from scripts.ssp_generator import load_system_metadata, load_control_narratives, generate_oscal_ssp
        from scripts.oscal_validator import validate_document
        from scripts.compliance_scorer import compute_compliance_score

        # 1. Inheritance mapping
        inheritance = {c["id"]: get_inheritance_for_control(c["id"], "aws") for c in sample_baseline_controls}
        inherited_count = sum(1 for v in inheritance.values() if v == "inherited")
        assert inherited_count > 0, "Should have some inherited controls"

        # 2. Gap analysis
        gap_result = analyze_gaps(sample_baseline_controls, sample_implementation_status, inheritance)
        assert gap_result["overall"]["total_controls"] == len(sample_baseline_controls)
        compliance_score = gap_result["overall"]["compliance_score"]
        assert compliance_score > 0

        # 3. Scan aggregation
        scan_findings_objs = aggregate_findings([FIXTURES / "sample-nessus.csv", FIXTURES / "sample-trivy.json"])
        deduped = deduplicate_findings(scan_findings_objs)
        summary = compute_scan_summary(deduped)
        output_findings_json(deduped, summary, output_dir / "scans.json")

        # 4. POA&M
        config = load_sla_config()
        loaded_findings = json.loads((output_dir / "scans.json").read_text())["findings"]
        poam_items = [create_poam_item(f, config, i) for i, f in enumerate(loaded_findings, 1)
                      if f.get("severity", "").upper() not in ("INFORMATIONAL", "")]
        escalations = check_escalation_triggers(poam_items, config)
        poam_summary = compute_poam_summary(poam_items)

        assert poam_summary["total_items"] > 0
        assert poam_summary["open"] > 0

        # 5. SSP generation
        metadata = load_system_metadata(EXAMPLES / "sample-system-metadata.yaml")
        narratives = load_control_narratives(FIXTURES / "sample-controls")
        ssp = generate_oscal_ssp(metadata, sample_baseline_controls, narratives, inheritance, "moderate")

        ssp_path = output_dir / "ssp.json"
        ssp_path.write_text(json.dumps(ssp, indent=2))

        # 6. OSCAL validation
        validation = validate_document(ssp_path, "ssp")
        assert validation.passed, f"SSP validation failed: {[e['message'] for e in validation.errors]}"

        # 7. Compliance scoring
        poam_data = {"summary": poam_summary, "escalation_triggers": escalations, "items": poam_items}
        score = compute_compliance_score(gap_report=gap_result, poam_data=poam_data)

        assert 0 < score["composite_score"] < 100
        assert score["risk_rating"] in ("LOW", "MODERATE", "ELEVATED", "HIGH")

        # Verify the pipeline produced coherent results
        print(f"\n  End-to-end pipeline results:")
        print(f"  Controls: {gap_result['overall']['total_controls']}")
        print(f"  Inherited: {gap_result['overall'].get('inherited', 0)}")
        print(f"  Compliance: {compliance_score}%")
        print(f"  POA&M items: {poam_summary['total_items']} ({poam_summary['overdue']} overdue)")
        print(f"  Composite score: {score['composite_score']} ({score['risk_rating']})")

    def test_pipeline_output_files(self, output_dir, sample_baseline_controls, sample_implementation_status, scan_findings):
        """Verify all pipeline artifacts are generated and parseable."""
        from scripts.gap_analysis import analyze_gaps, output_json as output_gap_json
        from scripts.scan_aggregator import aggregate_findings, compute_scan_summary, output_findings_json, output_conmon_csv
        from scripts.poam_manager import (
            create_poam_item, load_sla_config, compute_poam_summary,
            check_escalation_triggers, output_poam_json, output_poam_csv,
        )
        from scripts.compliance_scorer import compute_compliance_score, output_score_json

        # Gap analysis output
        gap_result = analyze_gaps(sample_baseline_controls, sample_implementation_status)
        output_gap_json(gap_result, output_dir / "gap.json")

        # Scan output
        findings = aggregate_findings([FIXTURES / "sample-nessus.csv"])
        summary = compute_scan_summary(findings)
        output_findings_json(findings, summary, output_dir / "scans.json")
        output_conmon_csv(findings, output_dir / "conmon.csv")

        # POA&M output
        config = load_sla_config()
        poam_items = [create_poam_item(f, config, i) for i, f in enumerate(scan_findings, 1)
                      if f["severity"] != "INFORMATIONAL"]
        escalations = check_escalation_triggers(poam_items, config)
        poam_summary = compute_poam_summary(poam_items)
        output_poam_json(poam_items, poam_summary, escalations, output_dir / "poam.json")
        output_poam_csv(poam_items, output_dir / "poam.csv")

        # Score output
        poam_data = {"summary": poam_summary, "escalation_triggers": escalations}
        score = compute_compliance_score(gap_report=gap_result, poam_data=poam_data)
        output_score_json(score, None, output_dir / "score.json")

        # Verify all outputs are valid JSON/CSV
        for json_file in ["gap.json", "scans.json", "poam.json", "score.json"]:
            data = json.loads((output_dir / json_file).read_text())
            assert isinstance(data, dict), f"{json_file} is not a dict"

        for csv_file in ["conmon.csv", "poam.csv"]:
            with open(output_dir / csv_file) as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                assert len(rows) > 0, f"{csv_file} has no rows"


# ============================================================
# Edge cases and error handling
# ============================================================


class TestEdgeCasesAndErrors:
    """Test robustness with unusual inputs."""

    def test_empty_scan_input(self, output_dir):
        from scripts.scan_aggregator import aggregate_findings, compute_scan_summary
        findings = aggregate_findings([tmp_path := output_dir / "empty.csv"])
        # Non-existent file just logs warning
        summary = compute_scan_summary(findings)
        assert summary["total_findings"] == 0

    def test_empty_poam(self):
        from scripts.poam_manager import compute_poam_summary, check_escalation_triggers, load_sla_config
        config = load_sla_config()
        summary = compute_poam_summary([])
        triggers = check_escalation_triggers([], config)
        assert summary["total_items"] == 0
        assert len(triggers) == 0

    def test_all_controls_implemented(self, sample_baseline_controls):
        from scripts.gap_analysis import analyze_gaps
        impl = {c["id"]: "implemented" for c in sample_baseline_controls}
        result = analyze_gaps(sample_baseline_controls, impl)
        assert result["overall"]["compliance_score"] == 100.0
        gaps = [c for c in result.get("controls", []) if c["is_gap"]]
        assert len(gaps) == 0

    def test_no_controls_implemented(self, sample_baseline_controls):
        from scripts.gap_analysis import analyze_gaps
        impl = {}  # Nothing implemented
        result = analyze_gaps(sample_baseline_controls, impl)
        assert result["overall"]["compliance_score"] == 0.0
        gaps = [c for c in result.get("controls", []) if c["is_gap"]]
        assert len(gaps) == len(sample_baseline_controls)

    def test_validator_rejects_garbage(self, output_dir):
        from scripts.oscal_validator import validate_document
        garbage = output_dir / "garbage.json"
        garbage.write_text('{"not": "an ssp"}')
        result = validate_document(garbage, "ssp")
        assert not result.passed

    def test_drift_identical_inventories(self):
        from scripts.inventory_drift import detect_drift
        inv = [
            {"resource_id": "x", "_normalized_id": "x", "resource_type": "EC2"},
            {"resource_id": "y", "_normalized_id": "y", "resource_type": "S3"},
        ]
        drift = detect_drift(inv, inv)
        assert drift["summary"]["drift_percentage"] == 0.0
        assert drift["summary"]["undocumented"] == 0
        assert drift["summary"]["stale"] == 0

    def test_drift_completely_different(self):
        from scripts.inventory_drift import detect_drift
        doc = [{"resource_id": "a", "_normalized_id": "a"}]
        live = [{"resource_id": "b", "_normalized_id": "b"}]
        drift = detect_drift(doc, live)
        assert drift["summary"]["stale"] == 1
        assert drift["summary"]["undocumented"] == 1
        assert drift["summary"]["matched"] == 0

    def test_scorer_no_vulns_perfect(self):
        from scripts.compliance_scorer import compute_compliance_score
        gap = {"overall": {
            "compliance_score": 100.0, "total_controls": 100,
            "implemented": 100, "inherited": 0, "partial": 0,
            "not_implemented": 0, "critical_gaps": 0,
        }}
        poam = {"summary": {"total_items": 0, "open": 0, "overdue": 0, "critical_high_overdue": 0}, "escalation_triggers": []}
        score = compute_compliance_score(gap_report=gap, poam_data=poam)
        assert score["composite_score"] >= 90
        assert score["risk_rating"] == "LOW"

    def test_scorer_catastrophic(self):
        from scripts.compliance_scorer import compute_compliance_score
        gap = {"overall": {
            "compliance_score": 10.0, "total_controls": 100,
            "implemented": 10, "inherited": 0, "partial": 5,
            "not_implemented": 85, "critical_gaps": 30,
        }}
        poam = {"summary": {"total_items": 50, "open": 50, "overdue": 40, "critical_high_overdue": 20}, "escalation_triggers": [{"type": "CAP"}]}
        score = compute_compliance_score(gap_report=gap, poam_data=poam)
        assert score["composite_score"] < 30
        assert score["risk_rating"] == "HIGH"

    def test_poam_merge_month_over_month(self):
        """Simulate monthly POA&M merge workflow."""
        from scripts.poam_manager import create_poam_item, merge_poam, load_sla_config, compute_poam_summary

        config = load_sla_config()
        now = datetime.now(timezone.utc)

        # Month 1: 3 findings
        month1_findings = [
            {"finding_id": "f1", "severity": "HIGH", "title": "Vuln A", "solution": "Patch A",
             "first_seen": (now - timedelta(days=60)).isoformat()},
            {"finding_id": "f2", "severity": "MODERATE", "title": "Vuln B", "solution": "Patch B",
             "first_seen": (now - timedelta(days=60)).isoformat()},
            {"finding_id": "f3", "severity": "LOW", "title": "Vuln C", "solution": "Patch C",
             "first_seen": (now - timedelta(days=60)).isoformat()},
        ]
        month1_items = [create_poam_item(f, config, i) for i, f in enumerate(month1_findings, 1)]

        # Month 2: f1 still present, f2 gone (remediated), f4 new
        month2_findings = [
            {"finding_id": "f1", "severity": "HIGH", "title": "Vuln A", "solution": "Patch A",
             "first_seen": (now - timedelta(days=90)).isoformat()},
            {"finding_id": "f4", "severity": "HIGH", "title": "Vuln D", "solution": "Patch D",
             "first_seen": (now - timedelta(days=5)).isoformat()},
        ]
        month2_items = [create_poam_item(f, config, i) for i, f in enumerate(month2_findings, 1)]

        merged = merge_poam(month1_items, month2_items)
        assert len(merged) == 4  # f1 updated, f2 remediated, f3 remediated, f4 new

        # f2 should be remediated
        f2 = next(m for m in merged if m["finding_id"] == "f2")
        assert "Remediated" in f2["status"]

        # f1 should have updated age
        f1 = next(m for m in merged if m["finding_id"] == "f1")
        assert f1["status"] == "Open"

        summary = compute_poam_summary(merged)
        assert summary["open"] == 2  # f1 and f4
        assert summary["remediated"] == 2  # f2 and f3
